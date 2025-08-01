import streamlit as st 
import pandas as pd 
import numpy as np 
import matplotlib.pyplot as plt 
from io import BytesIO
from openpyxl import load_workbook

# 设置页面配置
st.set_page_config(page_title="Excel数据处理平台", layout="wide")
st.title("Excel数据处理平台")
st.caption("作者：周福来")

# 自定义CSS样式优化显示
st.markdown(""" 
<style>
    .功能块 {border: 1px solid #e0e0e0; border-radius: 8px; padding: 15px; margin: 10px 0;}
    .模块标题 {font-size: 1.2rem; font-weight: bold; margin-bottom: 10px;}
    .stButton > button {width: 100%;}
    .column-block {border: 1px solid #e6f7ff; border-radius: 6px; padding: 10px; margin: 10px 0; background-color: #f0f7ff;}
    .column-header {font-weight: bold; color: #1890ff; margin-bottom: 8px;}
    .scrollable-container {max-height: 400px; overflow-y: auto; padding: 10px;}
    .merge-result {background-color: #f0fff4; border-left: 4px solid #52c41a; padding: 10px; margin: 10px 0;}
    .extract-result {background-color: #fff7e6; border-left: 4px solid #faad14; padding: 10px; margin: 10px 0;}
    .search-result {background-color: #e6f7ff; border-left: 4px solid #1890ff; padding: 10px; margin: 10px 0;}
    .compare-result {background-color: #fff1f0; border-left: 4px solid #f5222d; padding: 10px; margin: 10px 0;}
    .aircraft-match {background-color: #e6fffb; border-left: 4px solid #36cfc9; padding: 10px; margin: 10px 0;}
    .value-category {border-radius: 4px; padding: 5px; margin: 2px 0;}
    .common-value {background-color: #e6f7ff;}
    .only-a-value {background-color: #fff7e6;}
    .only-b-value {background-color: #f0fff4;}
    .matched-aircraft {background-color: #e6fffb;}
    .only-a-aircraft {background-color: #fff1f0;}
    .only-b-aircraft {background-color: #fff7e6;}
    /* 调整功能块内部元素间距 */
    .功能块 .stRadio, .功能块 .stTextInput, .功能块 .stSelectbox {margin-bottom: 10px;}
    /* 单元格选择表格样式 */
    .cell-selection-row {display: flex; width: 100%;}
    .row-selector {flex: 0 0 80px; padding: 5px;}
    .cell-column {flex: 1; padding: 5px; min-width: 120px;}
</style>
""", unsafe_allow_html=True)

# 辅助函数
def generate_key(prefix, row, col):
    return f"{prefix}_{row}_{col}"

def get_selected_cells(df, session_state_key):
    selected = []
    if session_state_key in st.session_state:
        for key, value in st.session_state[session_state_key].items():
            if value:
                parts = key.split("_")
                row = int(parts[1])
                col = int(parts[2])
                selected.append((row, col))
    return selected

def get_selected_columns(df, session_state_key):
    """获取选中列函数，准确识别整列选择"""
    if session_state_key not in st.session_state:
        return []
    
    # 统计每列被选中的单元格数量
    col_selection_count = {col: 0 for col in range(len(df.columns))}
    total_rows = len(df)
    
    for key, value in st.session_state[session_state_key].items():
        if value:
            parts = key.split("_")
            if len(parts) == 3 and parts[0] == "cell":
                try:
                    col = int(parts[2])
                    if 0 <= col < len(df.columns):
                        col_selection_count[col] += 1
                except (ValueError, IndexError):
                    continue  # 忽略无效的键
    
    # 判断是否整列被选中（超过80%的单元格被选中视为整列选择）
    selected_cols = []
    for col, count in col_selection_count.items():
        if count > 0 and (count / total_rows) > 0.8:
            selected_cols.append(col)
    
    return sorted(selected_cols)

def get_selected_rows(df, session_state_key):
    """获取选中行函数，准确识别整行选择"""
    if session_state_key not in st.session_state:
        return []
    
    # 统计每行被选中的单元格数量
    row_selection_count = {row: 0 for row in range(len(df))}
    total_cols = len(df.columns)
    
    for key, value in st.session_state[session_state_key].items():
        if value:
            parts = key.split("_")
            if len(parts) == 3 and parts[0] == "cell":
                try:
                    row = int(parts[1])
                    if 0 <= row < len(df):
                        row_selection_count[row] += 1
                except (ValueError, IndexError):
                    continue  # 忽略无效的键
    
    # 判断是否整行被选中（超过80%的单元格被选中视为整行选择）
    selected_rows = []
    for row, count in row_selection_count.items():
        if count > 0 and (count / total_cols) > 0.8:
            selected_rows.append(row)
    
    return sorted(selected_rows)

def set_row_selection(df, session_state_key, row, value):
    if session_state_key not in st.session_state:
        st.session_state[session_state_key] = {}
    for col in range(len(df.columns)):
        key = generate_key("cell", row, col)
        st.session_state[session_state_key][key] = value

def set_col_selection(df, session_state_key, col, value):
    """改进的设置列选择函数，确保所有单元格状态同步"""
    if session_state_key not in st.session_state:
        st.session_state[session_state_key] = {}
    # 遍历所有行，确保整列状态一致
    for row in range(len(df)):
        key = generate_key("cell", row, col)
        st.session_state[session_state_key][key] = value

def set_all_selection(df, session_state_key, value):
    """改进的全选函数，确保所有单元格状态同步"""
    if session_state_key not in st.session_state:
        st.session_state[session_state_key] = {}
    # 遍历所有行和列，确保所有单元格状态一致
    for row in range(len(df)):
        for col in range(len(df.columns)):
            key = generate_key("cell", row, col)
            st.session_state[session_state_key][key] = value

def compare_aircraft_lists(str_a, str_b, delimiter=","):
    """
    对比两个飞机列表字符串
    返回匹配的飞机、仅A有的飞机、仅B有的飞机
    """
    # 处理空值
    if pd.isna(str_a) or str_a.strip() == "":
        list_a = []
    else:
        # 分割、去空格、过滤空值
        list_a = [item.strip() for item in str_a.split(delimiter) if item.strip()]
    
    if pd.isna(str_b) or str_b.strip() == "":
        list_b = []
    else:
        list_b = [item.strip() for item in str_b.split(delimiter) if item.strip()]
    
    # 转换为集合进行比较
    set_a = set(list_a)
    set_b = set(list_b)
    
    # 计算匹配和不匹配的飞机
    matched = list(set_a & set_b)  # 匹配的飞机
    only_a = list(set_a - set_b)   # 仅A有的飞机
    only_b = list(set_b - set_a)   # 仅B有的飞机
    
    return {
        "matched": sorted(matched),
        "only_a": sorted(only_a),
        "only_b": sorted(only_b),
        "is_identical": len(only_a) == 0 and len(only_b) == 0  # 是否完全相同
    }

# 单个Excel处理模块
st.markdown("### 单个Excel处理模块")
uploaded_file = st.file_uploader("上传单个Excel文件", type=['xlsx', 'xls'], key="single_file")

if uploaded_file is not None:
    try:
        # 加载Excel文件获取所有sheet
        workbook = load_workbook(uploaded_file, read_only=True, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        # 重置文件时清除sheet选择状态
        if "last_uploaded_file" not in st.session_state or st.session_state.last_uploaded_file != uploaded_file.name:
            st.session_state.selected_sheet = sheet_names[0]
            st.session_state.last_uploaded_file = uploaded_file.name
        
        # Sheet选择器
        selected_sheet = st.selectbox(
            "选择当前工作表(SHEET)",
            sheet_names,
            index=sheet_names.index(st.session_state.selected_sheet),
            key="selected_sheet"
        )
        
        # 读取选中的sheet
        df = pd.read_excel(uploaded_file, sheet_name=selected_sheet, engine='openpyxl')
        st.success(f"文件上传成功！已加载工作表: {selected_sheet}")
        
        # 处理数据类型，确保Arrow序列化兼容
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    # 尝试转换为更合适的类型
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                except:
                    pass
        
        # 显示原始数据（整体展示）
        st.subheader(f"原始数据 - {selected_sheet}（整体展示）")
        st.markdown('<div class="scrollable-container">', unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True, height= None)
        st.markdown('</div>', unsafe_allow_html=True)
        
        # 单元格选择区域 - 为每个sheet创建独立的选择状态
        selection_key = f"cell_selections_{selected_sheet}"
        if selection_key not in st.session_state:
            st.session_state[selection_key] = {}
        
        # 整列选择区域（单独展示每列的选择框）
        st.subheader("整列选择")
        col_select_container = st.container()
        with col_select_container:
            # 按列显示选择框，每行显示5个列选择框
            cols_per_row = 5
            for i in range(0, len(df.columns), cols_per_row):
                row_cols = st.columns(cols_per_row)
                for j in range(cols_per_row):
                    col_idx = i + j
                    if col_idx >= len(df.columns):
                        break
                    col_name = df.columns[col_idx]
                    with row_cols[j]:
                        # 计算当前列的选中比例
                        total_rows = len(df)
                        selected_count = sum(
                            1 for row in range(len(df)) 
                            if st.session_state[selection_key].get(generate_key("cell", row, col_idx), False)
                        )
                        col_select_state = selected_count / total_rows > 0.8 if total_rows > 0 else False
                        
                        # 列选择回调函数
                        def col_callback(col_idx, selection_key=selection_key):
                            current_state = st.session_state[f"col_select_{selected_sheet}_{col_idx}"]
                            set_col_selection(df, selection_key, col_idx, current_state)
                        
                        col_select = st.checkbox(
                            f"{col_name}", 
                            key=f"col_select_{selected_sheet}_{col_idx}",
                            value=col_select_state,
                            on_change=col_callback,
                            args=(col_idx,)
                        )
        
        # 选择控制区
        col_select_all, col_select_info = st.columns([1, 3])
        with col_select_all:
            # 全选框状态同步
            all_selected = all(
                st.session_state[selection_key].get(generate_key("cell", row, col), False)
                for row in range(min(100, len(df)))  # 采样检查前100行
                for col in range(min(10, len(df.columns)))
            ) if len(df) > 0 and len(df.columns) > 0 else False
            
            # 全选回调函数
            def all_callback(selection_key=selection_key):
                current_state = st.session_state[f"select_all_{selected_sheet}"]
                set_all_selection(df, selection_key, current_state)
                # 更新列选择框状态
                for col_idx in range(len(df.columns)):
                    key = f"col_select_{selected_sheet}_{col_idx}"
                    if key in st.session_state:
                        st.session_state[key] = current_state
            
            select_all = st.checkbox(
                "全选所有单元格", 
                key=f"select_all_{selected_sheet}",
                value=all_selected,
                on_change=all_callback
            )
        with col_select_info:
            selected_cells = get_selected_cells(df, selection_key)
            selected_cols = get_selected_columns(df, selection_key)
            selected_rows = get_selected_rows(df, selection_key)
            st.info(f"已选中 {len(selected_cells)} 个单元格，{len(selected_rows)} 行，{len(selected_cols)} 列")
        
        # 分页显示带复选框的数据表格（增加滚动功能）
        max_display_rows = 15  # 每页显示15行数据
        total_rows = len(df)
        pages = (total_rows + max_display_rows - 1) // max_display_rows
        
        st.subheader("单元格选择（滚动浏览）")
        # 使用滚动容器包装分页内容
        st.markdown('<div class="scrollable-container">', unsafe_allow_html=True)
        
        # 页码滑动选择器
        col1, col2 = st.columns([3, 1])
        with col1:
            current_page = st.slider(
                "选择页码", 
                1, 
                pages, 
                1, 
                key=f"page_slider_{selected_sheet}",
                format="第 %d 页"
            )
        
        start_row = (current_page - 1) * max_display_rows
        end_row = min(start_row + max_display_rows, total_rows)
        with col2:
            st.write(f"显示行: {start_row + 1} - {end_row}")
        
        # 显示列标题行
        header_cols = st.columns([1] + [2]*min(10, len(df.columns)))
        with header_cols[0]:
            st.write("**行选择**")
        for col_idx, col_name in enumerate(df.columns[:10]):
            with header_cols[col_idx + 1]:
                st.write(f"**{col_name}**")
        
        # 显示当前页的单元格选择表格（修复布局问题）
        for row_idx in range(start_row, end_row):
            # 创建当前行的列布局
            row_cols = st.columns([1] + [2]*min(10, len(df.columns)))
            
            # 行选择复选框
            with row_cols[0]:
                # 行选择框状态同步
                row_selected = all(
                    st.session_state[selection_key].get(generate_key("cell", row_idx, col), False)
                    for col in range(len(df.columns))
                )
                
                # 行选择回调函数
                def make_row_callback(row_idx, selection_key):
                    def callback():
                        current_state = st.session_state[f"select_row_{selected_sheet}_{row_idx}"]
                        set_row_selection(df, selection_key, row_idx, current_state)
                    return callback
                
                row_select = st.checkbox(
                    f"行 {row_idx}", 
                    key=f"select_row_{selected_sheet}_{row_idx}",
                    value=row_selected,
                    on_change=make_row_callback(row_idx, selection_key)
                )
            
            # 单元格内容和选择框
            for col_idx, col_name in enumerate(df.columns[:10]):
                with row_cols[col_idx + 1]:
                    cell_value = df.iloc[row_idx, col_idx]
                    cell_key = generate_key("cell", row_idx, col_idx)
                    # 初始化单元格状态
                    if cell_key not in st.session_state[selection_key]:
                        st.session_state[selection_key][cell_key] = False
                    
                    # 单元格选择回调函数
                    def make_cell_callback(row_idx, col_idx, selection_key):
                        def callback():
                            key = generate_key("cell", row_idx, col_idx)
                            current_state = st.session_state[f"{selected_sheet}_{key}"]
                            st.session_state[selection_key][key] = current_state
                            # 更新列选择框状态
                            total_rows = len(df)
                            selected_count = sum(
                                1 for r in range(len(df)) 
                                if st.session_state[selection_key].get(generate_key("cell", r, col_idx), False)
                            )
                            new_col_state = selected_count / total_rows > 0.8 if total_rows > 0 else False
                            col_key = f"col_select_{selected_sheet}_{col_idx}"
                            if col_key in st.session_state:
                                st.session_state[col_key] = new_col_state
                        return callback
                    
                    # 单元格选择复选框
                    checked = st.checkbox(
                        f"{col_name}", 
                        value=st.session_state[selection_key][cell_key], 
                        key=f"{selected_sheet}_{cell_key}", 
                        label_visibility="collapsed",
                        on_change=make_cell_callback(row_idx, col_idx, selection_key)
                    )
                    st.session_state[selection_key][cell_key] = checked
                    st.write(str(cell_value) if pd.notna(cell_value) else "空")
        
        if len(df.columns) > 10:
            st.info(f"仅显示前10列，共 {len(df.columns)} 列")
        
        st.markdown('</div>', unsafe_allow_html=True)  # 关闭滚动容器
        
        # 核心功能区（同一行显示五个功能）
        st.subheader("数据处理功能")
        func_cols = st.columns(5)  # 一行五列布局
        
        # 1. 合并单元格
        with func_cols[0]:
            st.markdown('<div class="功能块">', unsafe_allow_html=True)
            st.markdown('<div class="模块标题">合并单元格</div>', unsafe_allow_html=True)
            st.write("将选中内容按指定维度合并")
            
            merge_dimension = st.radio(
                "合并维度",
                ["所有选中单元格合并为一个", "按行合并（每行一个结果）", "按列合并（每列一个结果）"],
                key=f"merge_dimension_{selected_sheet}",
                horizontal=True  # 水平排列单选框节省空间
            )
            
            separator = st.text_input("合并分隔符", value=", ", key=f"merge_separator_{selected_sheet}")
            
            if st.button("执行合并", key=f"merge_btn_{selected_sheet}"):
                selected_cells = get_selected_cells(df, selection_key)
                selected_rows = get_selected_rows(df, selection_key)
                selected_cols = get_selected_columns(df, selection_key)
                
                if not selected_cells and not selected_rows and not selected_cols:
                    st.warning("请先选择单元格、行或列")
                else:
                    if merge_dimension == "所有选中单元格合并为一个":
                        merge_data = []
                        for idx, (row, col) in enumerate(selected_cells, 1):
                            col_name = df.columns[col]
                            cell_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                            merge_data.append({
                                "序号": idx,
                                "位置": f"行{row},列{col_name}",
                                "内容": cell_value
                            })
                        merged = separator.join([item["内容"] for item in merge_data if item["内容"]])
                        
                        st.success("所有选中单元格合并完成")
                        st.dataframe(pd.DataFrame(merge_data), height=150)
                        st.markdown('<div class="merge-result">', unsafe_allow_html=True)
                        st.write(f"合并结果：{merged}")
                        st.markdown('</div>', unsafe_allow_html=True)
                    elif merge_dimension == "按行合并（每行一个结果）":
                        if not selected_rows:
                            rows_in_selected = sorted(list(set(row for row, col in selected_cells)))
                        else:
                            rows_in_selected = selected_rows
                        
                        row_merge_results = []
                        for row in rows_in_selected:
                            row_cells = [(r, c) for r, c in selected_cells if r == row]
                            if not row_cells:
                                continue
                            
                            row_cells.sort(key=lambda x: x[1])
                            cell_values = [str(df.iloc[r, c]) if pd.notna(df.iloc[r, c]) else "" for r, c in row_cells]
                            merged_row = separator.join(cell_values)
                            row_merge_results.append({
                                "行号": row,
                                "合并结果": merged_row,
                                "包含单元格数": len(row_cells)
                            })
                        
                        st.success(f"按行合并完成，共合并 {len(row_merge_results)} 行")
                        st.dataframe(pd.DataFrame(row_merge_results), height=200)
                    else:
                        if not selected_cols:
                            cols_in_selected = sorted(list(set(col for row, col in selected_cells)))
                        else:
                            cols_in_selected = selected_cols
                        
                        col_merge_results = []
                        for col in cols_in_selected:
                            col_cells = [(r, c) for r, c in selected_cells if c == col]
                            if not col_cells:
                                continue
                            
                            col_cells.sort(key=lambda x: x[0])
                            cell_values = [str(df.iloc[r, c]) if pd.notna(df.iloc[r, c]) else "" for r, c in col_cells]
                            merged_col = separator.join(cell_values)
                            col_merge_results.append({
                                "列名": df.columns[col],
                                "列索引": col,
                                "合并结果": merged_col,
                                "包含单元格数": len(col_cells)
                            })
                        
                        st.success(f"按列合并完成，共合并 {len(col_merge_results)} 列")
                        st.dataframe(pd.DataFrame(col_merge_results), height=200)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # 2. 拆分单元格
        with func_cols[1]:
            st.markdown('<div class="功能块">', unsafe_allow_html=True)
            st.markdown('<div class="模块标题">拆分单元格</div>', unsafe_allow_html=True)
            st.write("用分隔符拆分选中单元格内容")
            delimiter = st.text_input("分隔符", key=f"delimiter_{selected_sheet}", placeholder=",")
            
            # 添加拆分选项
            split_option = st.radio(
                "拆分选项",
                ["拆分后保留空值", "拆分后过滤空值"],
                key=f"split_option_{selected_sheet}",
                horizontal=True
            )
            
            if st.button("执行拆分", key=f"split_btn_{selected_sheet}"):
                selected = get_selected_cells(df, selection_key)
                if not selected:
                    st.warning("请选择单元格")
                else:
                    for idx, (row, col) in enumerate(selected, 1):
                        col_name = df.columns[col]
                        value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                        st.write(f"单元格 {idx}：行{row},列{col_name}")
                        if not delimiter:
                            st.warning("请输入分隔符")
                            break
                        if delimiter not in value and value:
                            st.warning(f"无'{delimiter}'分隔符")
                        else:
                            split_res = [v.strip() for v in value.split(delimiter)] if value else []
                            # 根据选项过滤空值
                            if split_option == "拆分后过滤空值":
                                split_res = [v for v in split_res if v]
                            st.dataframe(pd.DataFrame({
                                "拆分序号": range(1, len(split_res)+1),
                                "内容": split_res
                            }), height=150)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # 3. 对比两列
        with func_cols[2]:
            st.markdown('<div class="功能块">', unsafe_allow_html=True)
            st.markdown('<div class="模块标题">对比两列</div>', unsafe_allow_html=True)
            st.write("对比两列数据差异")
            
            # 列选择
            col1 = st.selectbox("选择列A", df.columns, key=f"col_a_{selected_sheet}")
            col2 = st.selectbox("选择列B", df.columns, key=f"col_b_{selected_sheet}")
            
            # 对比类型选择
            compare_type = st.radio(
                "对比类型",
                ["普通对比", "飞机对比"],
                key=f"compare_type_{selected_sheet}",
                horizontal=True
            )
            
            # 飞机对比的分隔符设置（仅在选择飞机对比时显示）
            if compare_type == "飞机对比":
                aircraft_delimiter = st.text_input(
                    "飞机编号分隔符", 
                    value=",", 
                    key=f"aircraft_delimiter_{selected_sheet}",
                    max_chars=2,
                    help="用于分割飞机编号的字符，如逗号、分号等"
                )
            
            # 对比选项
            compare_option = st.radio(
                "对比选项",
                ["完整对比", "仅显示差异"],
                key=f"compare_option_{selected_sheet}",
                horizontal=True
            )
            
            # 显示值类别选项
            show_values_option = st.checkbox(
                "显示类型详情表",
                key=f"show_values_option_{selected_sheet}",
                value=True
            )
            
            if st.button("执行对比", key=f"compare_btn_{selected_sheet}"):
                if col1 == col2:
                    st.warning("请选择不同列进行对比")
                else:
                    # 普通对比逻辑
                    if compare_type == "普通对比":
                        # 准备对比数据
                        data_a = [str(x) if pd.notna(x) else "" for x in df[col1]]
                        data_b = [str(x) if pd.notna(x) else "" for x in df[col2]]
                        
                        # 计算差异
                        diff_indices = [i for i in range(len(data_a)) if i < len(data_b) and data_a[i] != data_b[i]]
                        
                        # 统计
                        total = min(len(data_a), len(data_b))
                        same = total - len(diff_indices)
                        diff_percent = (len(diff_indices) / total * 100) if total > 0 else 0
                        
                        st.success(f"对比完成：共 {total} 行数据，相同 {same} 行，差异 {len(diff_indices)} 行（差异率 {diff_percent:.2f}%）")
                        
                        # 计算值类别
                        set_a = set([x for x in data_a if x])
                        set_b = set([x for x in data_b if x])
                        common_values = list(set_a & set_b)
                        only_a_values = list(set_a - set_b)
                        only_b_values = list(set_b - set_a)
                        
                        # 整合为类型详情表（类型、数量、明细）
                        detail_data = [
                            {
                                "类型": f"{col1}与{col2}共有值",
                                "数量": len(common_values),
                                "明细": ", ".join(common_values[:50]) + ("..." if len(common_values) > 50 else "")
                            },
                            {
                                "类型": f"仅{col1}有",
                                "数量": len(only_a_values),
                                "明细": ", ".join(only_a_values[:50]) + ("..." if len(only_a_values) > 50 else "")
                            },
                            {
                                "类型": f"仅{col2}有",
                                "数量": len(only_b_values),
                                "明细": ", ".join(only_b_values[:50]) + ("..." if len(only_b_values) > 50 else "")
                            }
                        ]
                        detail_df = pd.DataFrame(detail_data)
                        
                        # 显示类型详情表
                        if show_values_option:
                            st.subheader("类型详情汇总表")
                            st.dataframe(detail_df, use_container_width=True)
                        
                        # 详细对比结果
                        st.markdown('<div class="compare-result">', unsafe_allow_html=True)
                        if compare_option == "完整对比" or len(diff_indices) <= 100:
                            compare_results = []
                            for i in range(min(len(data_a), len(data_b))):
                                status = "相同" if data_a[i] == data_b[i] else "不同"
                                if compare_option == "完整对比" or status == "不同":
                                    compare_results.append({
                                        "行号": i,
                                        col1: data_a[i],
                                        col2: data_b[i],
                                        "状态": status
                                    })
                            st.dataframe(pd.DataFrame(compare_results), height=200)
                        else:
                            st.info(f"差异数据较多（{len(diff_indices)}行），仅显示前100行差异")
                            compare_results = []
                            count = 0
                            for i in range(min(len(data_a), len(data_b))):
                                if data_a[i] != data_b[i]:
                                    compare_results.append({
                                        "行号": i,
                                        col1: data_a[i],
                                        col2: data_b[i],
                                        "状态": "不同"
                                    })
                                    count += 1
                                    if count >= 100:
                                        break
                            st.dataframe(pd.DataFrame(compare_results), height=200)
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    # 飞机对比逻辑
                    else:  # compare_type == "飞机对比"
                        # 获取分隔符
                        delimiter = aircraft_delimiter if 'aircraft_delimiter' in locals() else ","
                        
                        # 存储所有行的对比结果
                        all_results = []
                        # 累计所有匹配和不匹配的飞机
                        total_matched = set()
                        total_only_a = set()
                        total_only_b = set()
                        
                        # 逐行对比
                        for row_idx in range(min(len(df[col1]), len(df[col2]))):
                            val_a = df.iloc[row_idx][col1]
                            val_b = df.iloc[row_idx][col2]
                            
                            # 转换为字符串进行处理
                            str_a = str(val_a) if pd.notna(val_a) else ""
                            str_b = str(val_b) if pd.notna(val_b) else ""
                            
                            # 对比飞机列表
                            result = compare_aircraft_lists(str_a, str_b, delimiter)
                            
                            # 更新累计集合
                            total_matched.update(result["matched"])
                            total_only_a.update(result["only_a"])
                            total_only_b.update(result["only_b"])
                            
                            # 记录结果
                            all_results.append({
                                "行号": row_idx,
                                f"{col1}原始值": str_a,
                                f"{col2}原始值": str_b,
                                "匹配的飞机": ", ".join(result["matched"]) if result["matched"] else "无",
                                f"仅{col1}有的飞机": ", ".join(result["only_a"]) if result["only_a"] else "无",
                                f"仅{col2}有的飞机": ", ".join(result["only_b"]) if result["only_b"] else "无",
                                "状态": "完全匹配" if result["is_identical"] else "不完全匹配"
                            })
                        
                        # 整合为飞机类型详情表（类型、数量、明细）
                        detail_data = [
                            {
                                "类型": "所有匹配的飞机",
                                "数量": len(total_matched),
                                "明细": ", ".join(sorted(total_matched)[:50]) + ("..." if len(total_matched) > 50 else "")
                            },
                            {
                                "类型": f"仅{col1}有的飞机",
                                "数量": len(total_only_a),
                                "明细": ", ".join(sorted(total_only_a)[:50]) + ("..." if len(total_only_a) > 50 else "")
                            },
                            {
                                "类型": f"仅{col2}有的飞机",
                                "数量": len(total_only_b),
                                "明细": ", ".join(sorted(total_only_b)[:50]) + ("..." if len(total_only_b) > 50 else "")
                            }
                        ]
                        detail_df = pd.DataFrame(detail_data)
                        
                        # 显示总体统计
                        st.success(
                            f"飞机对比完成：共 {len(all_results)} 行数据，"
                            f"完全匹配 {sum(1 for r in all_results if r['状态'] == '完全匹配')} 行，"
                            f"不完全匹配 {sum(1 for r in all_results if r['状态'] == '不完全匹配')} 行"
                        )
                        
                        # 显示类型详情表
                        if show_values_option:
                            st.subheader("飞机类型详情汇总表")
                            st.dataframe(detail_df, use_container_width=True)
                        
                        # 显示行级对比结果
                        st.subheader("行级飞机对比结果")
                        st.markdown('<div class="aircraft-match">', unsafe_allow_html=True)
                        
                        if compare_option == "完整对比" or sum(1 for r in all_results if r["状态"] == "不完全匹配") <= 100:
                            # 显示所有行或所有不匹配的行
                            display_results = [
                                r for r in all_results 
                                if compare_option == "完整对比" or r["状态"] == "不完全匹配"
                            ]
                            st.dataframe(pd.DataFrame(display_results), height=300)
                        else:
                            # 仅显示前100行不匹配的结果
                            st.info(f"不完全匹配的数据较多，仅显示前100行")
                            display_results = []
                            count = 0
                            for r in all_results:
                                if r["状态"] == "不完全匹配":
                                    display_results.append(r)
                                    count += 1
                                    if count >= 100:
                                        break
                            st.dataframe(pd.DataFrame(display_results), height=300)
                        
                        st.markdown('</div>', unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        # 4. 数据提取模块
        with func_cols[3]:
            st.markdown('<div class="功能块">', unsafe_allow_html=True)
            st.markdown('<div class="模块标题">数据提取</div>', unsafe_allow_html=True)
            st.write("根据特定字符提取内容")
            
            extract_type = st.radio(
                "提取类型",
                ["左侧数据", "右侧数据", "中间数据"],
                key=f"extract_type_{selected_sheet}",
                horizontal=True
            )
            
            # 简化标签显示，节省空间
            if extract_type in ["左侧数据", "右侧数据"]:
                char = st.text_input("请输入分隔字符", key=f"single_char_{selected_sheet}", max_chars=1)
                extract_full_type = f"提取特定字符{extract_type}"
            else:
                char1 = st.text_input("第一个字符", key=f"char1_{selected_sheet}", max_chars=1)
                char2 = st.text_input("第二个字符", key=f"char2_{selected_sheet}", max_chars=1)
                extract_full_type = "提取两个特定字符中间数据"
            
            if st.button("执行提取", key=f"extract_btn_{selected_sheet}"):
                selected = get_selected_cells(df, selection_key)
                if not selected:
                    st.warning("请选择单元格")
                else:
                    if extract_type in ["左侧数据", "右侧数据"] and not char:
                        st.warning("请输入分隔字符")
                    elif extract_type == "中间数据" and (not char1 or not char2):
                        st.warning("请输入两个字符")
                    else:
                        results = []
                        for idx, (row, col) in enumerate(selected, 1):
                            col_name = df.columns[col]
                            value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                            
                            if not value:
                                results.append({
                                    "序号": idx,
                                    "位置": f"行{row},列{col_name}",
                                    "原始内容": value,
                                    "提取结果": "（空值）"
                                })
                                continue
                            
                            extracted = ""
                            if extract_type == "左侧数据":
                                if char in value:
                                    extracted = value.split(char)[0]
                                else:
                                    extracted = f"（无'{char}'字符）"
                            elif extract_type == "右侧数据":
                                if char in value:
                                    extracted = char.join(value.split(char)[1:])
                                else:
                                    extracted = f"（无'{char}'字符）"
                            else:  # 中间数据
                                if char1 in value and char2 in value:
                                    start_idx = value.find(char1) + 1
                                    end_idx = value.find(char2, start_idx)
                                    if end_idx != -1:
                                        extracted = value[start_idx:end_idx]
                                    else:
                                        extracted = f"（'{char1}'后无'{char2}'）"
                                else:
                                    missing = []
                                    if char1 not in value:
                                        missing.append(f"'{char1}'")
                                    if char2 not in value:
                                        missing.append(f"'{char2}'")
                                    extracted = f"（无{', '.join(missing)}字符）"
                            
                            results.append({
                                "序号": idx,
                                "位置": f"行{row},列{col_name}",
                                "原始内容": value,
                                "提取结果": extracted
                            })
                        
                        st.success(f"完成 {len(results)} 个单元格的提取")
                        st.dataframe(pd.DataFrame(results), height=200)
                        
                        st.markdown('<div class="extract-result">', unsafe_allow_html=True)
                        st.write("提取结果汇总：")
                        for res in results:
                            st.write(f"{res['位置']}：{res['提取结果']}")
                        st.markdown('</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # 5. 数据查找模块
        with func_cols[4]:
            st.markdown('<div class="功能块">', unsafe_allow_html=True)
            st.markdown('<div class="模块标题">数据查找</div>', unsafe_allow_html=True)
            st.write("在其他工作表中查找选中内容")
            
            # 选择要查找的工作表
            other_sheets = [sheet for sheet in sheet_names if sheet != selected_sheet]
            
            if not other_sheets:
                st.warning("当前Excel文件只有一个工作表，无法进行跨工作表查找")
                target_sheet = None
            else:
                target_sheet = st.selectbox(
                    "目标工作表",
                    other_sheets,
                    key=f"target_sheet_{selected_sheet}"
                )
                
                # 读取目标工作表获取列信息
                target_df = pd.read_excel(uploaded_file, sheet_name=target_sheet, engine='openpyxl')
                target_col = st.selectbox(
                    "目标查找列",
                    target_df.columns,
                    key=f"target_col_{selected_sheet}"
                )
                
                # 查找选项
                search_option = st.radio(
                    "查找方式",
                    ["精确匹配", "包含匹配"],
                    key=f"search_option_{selected_sheet}",
                    horizontal=True
                )
            
            if st.button("执行查找", key=f"search_btn_{selected_sheet}"):
                selected = get_selected_cells(df, selection_key)
                if not selected:
                    st.warning("请先选择单元格")
                elif not other_sheets:
                    st.warning("没有可用于查找的其他工作表")
                elif not target_sheet or ('target_col' not in locals() and other_sheets):
                    st.warning("请完成目标工作表和列的选择")
                else:
                    # 读取目标工作表数据
                    target_df = pd.read_excel(uploaded_file, sheet_name=target_sheet, engine='openpyxl')
                    
                    # 对每个选中的单元格执行查找
                    for idx, (row, col) in enumerate(selected, 1):
                        col_name = df.columns[col]
                        search_value = str(df.iloc[row, col]) if pd.notna(df.iloc[row, col]) else ""
                        
                        st.markdown(f"### 查找结果 {idx}")
                        st.write(f"查找内容：{search_value}（位置：行{row},列{col_name}）")
                        st.write(f"查找范围：{target_sheet} 工作表的 [{target_col}] 列")
                        
                        if not search_value:
                            st.info("查找内容为空，无法执行查找")
                            continue
                        
                        # 根据选择执行不同查找方式
                        if search_option == "精确匹配":
                            matches = target_df[target_df[target_col].astype(str) == search_value]
                        else:  # 包含匹配
                            matches = target_df[target_df[target_col].astype(str).str.contains(search_value, na=False)]
                        
                        if len(matches) > 0:
                            st.success(f"找到 {len(matches)} 个匹配项")
                            st.markdown('<div class="search-result">', unsafe_allow_html=True)
                            st.dataframe(matches, height=200, use_container_width=True)
                            st.markdown('</div>', unsafe_allow_html=True)
                        else:
                            st.info(f"未找到与 '{search_value}' 匹配的内容")
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    except Exception as e:
        st.error(f"处理错误：{str(e)}")

# 多个Excel文件处理模块
st.markdown("### 多个Excel文件处理")
uploaded_files = st.file_uploader(
    "上传多个Excel文件（最多4个）", 
    type=['xlsx', 'xls'], 
    accept_multiple_files=True, 
    key="multi_files"
)

if uploaded_files:
    try:
        # 存储每个文件的所有sheet数据
        file_sheets = {}
        for f in uploaded_files[:4]:
            workbook = load_workbook(f, read_only=True, data_only=True)
            sheets = workbook.sheetnames
            workbook.close()
            file_sheets[f.name] = sheets
        
        # 选择要处理的文件
        selected_file = st.selectbox(
            "选择文件",
            list(file_sheets.keys()),
            key="multi_selected_file"
        )
        
        # 选择该文件的sheet
        selected_sheet = st.selectbox(
            "选择工作表(SHEET)",
            file_sheets[selected_file],
            key="multi_selected_sheet"
        )
        
        # 读取选中的文件和sheet
        df = pd.read_excel(
            uploaded_files[[f.name for f in uploaded_files].index(selected_file)],
            sheet_name=selected_sheet,
            engine='openpyxl'
        )
        
        st.success(f"已加载文件: {selected_file}，工作表: {selected_sheet}")
        
        # 处理数据类型
        for col in df.columns:
            if df[col].dtype == 'object':
                try:
                    df[col] = pd.to_numeric(df[col], errors='ignore')
                except:
                    pass
        
        # 显示原始数据（整体展示）
        st.subheader(f"原始数据 - {selected_file}[{selected_sheet}]（整体展示）")
        st.markdown('<div class="scrollable-container">', unsafe_allow_html=True)
        st.dataframe(df, use_container_width=True, height=None)
        st.markdown('</div>', unsafe_allow_html=True)
        
        file_info = []
        # 获取当前文件的所有sheet信息
        for sheet in file_sheets[selected_file]:
            temp_df = pd.read_excel(
                uploaded_files[[f.name for f in uploaded_files].index(selected_file)],
                sheet_name=sheet,
                engine='openpyxl'
            )
            col_names = [str(col) for col in temp_df.columns]
            display_cols = ", ".join(col_names[:5]) + (", ..." if len(col_names) > 5 else "")
            file_info.append({
                "工作表": sheet,
                "行数": len(temp_df),
                "列数": len(temp_df.columns),
                "前5列名": display_cols
            })
        
        st.subheader(f"{selected_file} 工作表信息汇总")
        st.dataframe(pd.DataFrame(file_info), use_container_width=True)
        
        if len(uploaded_files)>=2:
            st.subheader("跨文件对比功能")
            compare_cols = st.columns(4)
            
            with compare_cols[0]:
                file1 = st.selectbox("选择文件1", list(file_sheets.keys()), key="multi_file1")
                sheet1 = st.selectbox("选择工作表1", file_sheets[file1], key="multi_sheet1")
                df1 = pd.read_excel(
                    uploaded_files[[f.name for f in uploaded_files].index(file1)],
                    sheet_name=sheet1,
                    engine='openpyxl'
                )
            
            with compare_cols[1]:
                col1 = st.selectbox("选择列1", [str(col) for col in df1.columns], key="multi_col1")
                col1_original = df1.columns[df1.columns.astype(str) == col1][0]
            
            with compare_cols[2]:
                file2 = st.selectbox("选择文件2", [f for f in file_sheets.keys() if f != file1], key="multi_file2")
                sheet2 = st.selectbox("选择工作表2", file_sheets[file2], key="multi_sheet2")
                df2 = pd.read_excel(
                    uploaded_files[[f.name for f in uploaded_files].index(file2)],
                    sheet_name=sheet2,
                    engine='openpyxl'
                )
            
            with compare_cols[3]:
                col2 = st.selectbox("选择列2", [str(col) for col in df2.columns], key="multi_col2")
                col2_original = df2.columns[df2.columns.astype(str) == col2][0]
            
            if st.button("执行跨文件对比", key="multi_cross_compare_btn"):
                if col1 == col2 and file1 == file2 and sheet1 == sheet2:
                    st.warning("请选择不同的文件、工作表或列")
                else:
                    data1 = set(str(x) for x in df1[col1_original].dropna().unique())
                    data2 = set(str(x) for x in df2[col2_original].dropna().unique())
                    
                    common = list(data1 & data2)
                    only1 = list(data1 - data2)
                    only2 = list(data2 - data1)
                    
                    # 整合为类型详情表
                    detail_data = [
                        {
                            "类型": f"{file1}与{file2}共有值",
                            "数量": len(common),
                            "明细": ", ".join(common[:50]) + ("..." if len(common) > 50 else "")
                        },
                        {
                            "类型": f"仅{file1}有",
                            "数量": len(only1),
                            "明细": ", ".join(only1[:50]) + ("..." if len(only1) > 50 else "")
                        },
                        {
                            "类型": f"仅{file2}有",
                            "数量": len(only2),
                            "明细": ", ".join(only2[:50]) + ("..." if len(only2) > 50 else "")
                        }
                    ]
                    detail_df = pd.DataFrame(detail_data)
                    
                    st.success(f"对比 {file1}[{sheet1}!{col1}] 与 {file2}[{sheet2}!{col2}] 完成")
                    st.subheader("类型详情汇总表")
                    st.dataframe(detail_df, use_container_width=True)
                    
                    max_len = max(len(common), len(only1), len(only2))
                    detail_df = pd.DataFrame({
                        f"{file1}与{file2}共有": [common[i] if i < len(common) else "" for i in range(max_len)],
                        f"仅{file1}有": [only1[i] if i < len(only1) else "" for i in range(max_len)],
                        f"仅{file2}有": [only2[i] if i < len(only2) else "" for i in range(max_len)]
                    })
                    st.subheader("详细对比结果")
                    st.dataframe(detail_df, use_container_width=True)
        
        if len(uploaded_files)>=2:
            st.subheader("跨文件可视化")
            viz_cols = st.columns(3)
            with viz_cols[0]:
                viz_file1 = st.selectbox("选择文件1", list(file_sheets.keys()), key="viz_file1")
                viz_sheet1 = st.selectbox("选择工作表1", file_sheets[viz_file1], key="viz_sheet1")
                viz_df1 = pd.read_excel(
                    uploaded_files[[f.name for f in uploaded_files].index(viz_file1)],
                    sheet_name=viz_sheet1,
                    engine='openpyxl'
                )
            with viz_cols[1]:
                viz_col1 = st.selectbox("选择对比列", [str(col) for col in viz_df1.columns], key="viz_col1")
                viz_col1_original = viz_df1.columns[viz_df1.columns.astype(str) == viz_col1][0]
            with viz_cols[2]:
                viz_type = st.selectbox("图表类型", ["柱状图", "折线图", "饼图"], key="viz_type")
            
            if st.button("生成跨文件图表", key="cross_viz_btn"):
                num_data = []
                max_data_points = 200
                
                for f in uploaded_files[:4]:
                    first_sheet = file_sheets[f.name][0]
                    try:
                        df = pd.read_excel(f, sheet_name=first_sheet, engine='openpyxl')
                        if viz_col1_original in df.columns:
                            col_data = df[viz_col1_original]
                            valid_data = [val for val in col_data.dropna() if pd.api.types.is_numeric_dtype(type(val))]
                            take_count = min(len(valid_data), max_data_points // len(uploaded_files[:4]))
                            for val in valid_data[:take_count]:
                                num_data.append({"文件": f.name, "工作表": first_sheet, "数值": val})
                                if len(num_data) >= max_data_points:
                                    break
                    except Exception as e:
                        st.warning(f"处理文件 {f.name} 时出错: {str(e)}")
                        continue
                
                if num_data:
                    fig, ax = plt.subplots(figsize=(8, 5))
                    data_by_file = {}
                    for item in num_data:
                        key = f"{item['文件']}[{item['工作表']}]"
                        if key not in data_by_file:
                            data_by_file[key] = []
                        data_by_file[key].append(item["数值"])
                    
                    if viz_type == "柱状图":
                        x = np.arange(max(len(vals) for vals in data_by_file.values()))
                        width = 0.8 / len(data_by_file)
                        for i, (key, vals) in enumerate(data_by_file.items()):
                            ax.bar(
                                x + i*width, 
                                vals + [0]*(max(len(vals) for vals in data_by_file.values())-len(vals)),
                                width=width, 
                                label=key
                            )
                        ax.set_xticks(x + width*(len(data_by_file)-1)/2)
                        ax.set_xticklabels([f"数据点{i+1}" for i in range(x.size)])
                    elif viz_type == "折线图":
                        for key, vals in data_by_file.items():
                            ax.plot(range(1, len(vals)+1), vals, marker='o', label=key)
                        ax.set_xlabel("数据点索引")
                    elif viz_type == "饼图":
                        file_sum = {key: sum(vals) for key, vals in data_by_file.items()}
                        ax.pie(file_sum.values(), labels=file_sum.keys(), autopct='%1.1f%%')
                        ax.axis('equal')
                    
                    ax.set_ylabel("数值")
                    ax.legend(title="文件", bbox_to_anchor=(1.05, 1), loc='upper left')
                    plt.tight_layout()
                    st.pyplot(fig)
                else:
                    st.warning(f"未找到可可视化的数值数据（部分文件可能不含 {viz_col1} 列或该列无数值）")
    except Exception as e:
        st.error(f"处理错误：{str(e)}")